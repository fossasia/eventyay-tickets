import re
from collections import defaultdict
from datetime import timedelta
from io import BytesIO

import dateutil
import pytz
from django.core.files.base import ContentFile
from django.db.models import Q
from django.utils.timezone import is_naive, make_aware, now
from django.utils.translation import override
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from venueless.celery_app import app
from venueless.core.models import Channel, ExhibitorView, Room, RoomView
from venueless.core.models.world import WorldView
from venueless.core.tasks import WorldTask
from venueless.graphs.report import ReportGenerator
from venueless.graphs.utils import get_schedule, pretalx_uni18n
from venueless.storage.models import StoredFile


@app.task(base=WorldTask)
def generate_report(world, input=None):
    with override(world.locale):
        try:
            cf = ReportGenerator(world, pdf_graphs=True).build(input)
        except Exception:
            # We first try to embed graphs directly into the PDF for best quality. Sometimes that fails,
            # then we fall back to PNG rendering of graphs
            cf = ReportGenerator(world, pdf_graphs=False).build(input)
        return cf.file.url


@app.task(base=WorldTask)
def generate_attendee_list(world, input=None):
    io = BytesIO()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Attendees")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    for j in range(len(world.config.get("profile_fields", []))):
        ws.column_dimensions[get_column_letter(5 + j)].width = 30

    header = ["Internal ID", "External ID", "Name", "Permission traits"]
    for f in world.config.get("profile_fields", []):
        header.append(f.get("label"))
    ws.append(header)

    for u in world.user_set.all():
        ws.append(
            [
                str(u.pk),
                str(u.token_id) if u.token_id else "",
                u.profile.get("display_name") or "",
                ",".join(t for t in u.traits),
            ]
            + [
                u.profile.get("fields", {}).get(f.get("id") or f.get("label")) or ""
                for j, f in enumerate(world.config.get("profile_fields", []))
            ]
        )

    wb.save(io)
    io.seek(0)

    sf = StoredFile.objects.create(
        world=world,
        date=now(),
        filename="report.xlsx",
        expires=now() + timedelta(hours=2),
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        public=True,
        user=None,
    )
    sf.file.save("report.xlsx", ContentFile(io.read()))
    return sf.file.url


@app.task(base=WorldTask)
def generate_chat_history(world, input=None):
    channel = Channel.objects.get(pk=input.get("channel"))
    tz = pytz.timezone(world.timezone)
    io = BytesIO()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Messages")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 60

    header = ["Date", "Time", "Sender", "Content"]
    ws.append(header)

    for e in (
        channel.chat_events.filter(event_type="channel.message")
        .select_related("sender")
        .order_by("timestamp")
    ):
        if e.content.get("type") == "text":
            msg = e.content.get("body")
        elif e.content.get("type") == "files":
            msg = "\n".join(f.get("url") for f in e.content.get("files"))
            if e.content.get("body"):
                msg = e.content.get("body") + "\n" + msg
        else:
            msg = "<unknown message type>"
        ws.append(
            [
                e.timestamp.astimezone(tz).date(),
                e.timestamp.astimezone(tz).time(),
                e.sender.profile.get("display_name", ""),
                msg,
            ]
        )

    wb.save(io)
    io.seek(0)

    sf = StoredFile.objects.create(
        world=world,
        date=now(),
        filename="report.xlsx",
        expires=now() + timedelta(hours=2),
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        public=True,
        user=None,
    )
    sf.file.save("report.xlsx", ContentFile(io.read()))
    return sf.file.url


@app.task(base=WorldTask)
def generate_question_history(world, input=None):
    room = Room.objects.get(pk=input.get("room"))
    tz = pytz.timezone(world.timezone)
    io = BytesIO()

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Questions")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 40

    header = ["Date", "Time", "Content", "Votes", "Status"]
    ws.append(header)

    for e in room.questions.with_score().order_by("timestamp"):
        ws.append(
            [
                e.timestamp.astimezone(tz).date(),
                e.timestamp.astimezone(tz).time(),
                e.content,
                e.score,
                e.state,
            ]
        )

    wb.save(io)
    io.seek(0)

    sf = StoredFile.objects.create(
        world=world,
        date=now(),
        filename="report.xlsx",
        expires=now() + timedelta(hours=2),
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        public=True,
        user=None,
    )
    sf.file.save("report.xlsx", ContentFile(io.read()))
    return sf.file.url


@app.task(base=WorldTask)
def generate_room_views(world, input=None):
    wb = Workbook(write_only=True)
    io = BytesIO()
    tz = pytz.timezone(world.timezone)
    begin = dateutil.parser.parse(input.get("begin"))
    if is_naive(begin):
        make_aware(begin, tz)
    begin = begin.astimezone(tz)
    end = dateutil.parser.parse(input.get("end"))
    if is_naive(end):
        make_aware(end, tz)
    end = end.astimezone(tz)

    for room in world.rooms.all():
        types = [m["type"] for m in room.module_config]
        if any(
            t.startswith("livestream.")
            or t.startswith("chat.")
            or t.startswith("call.")
            for t in types
        ):
            name = re.sub("[^a-zA-Z0-9 ]", "", room.name)
            if room.deleted:
                name += " (deleted)"
            ws = wb.create_sheet(name)
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 20
            ws.append(["Date", "Time", "Viewership (approx. unique users)"])

            day = begin
            while day.date() <= end.date():
                gds = day.replace(hour=begin.hour, minute=0, second=0)
                gde = day.replace(hour=end.hour, minute=end.minute, second=0)

                views = (
                    room.views.filter(
                        Q(Q(end__isnull=True) | Q(end__gte=gds)) & Q(start__lte=gde)
                    )
                    .order_by()
                    .values("user", "start", "end")
                )

                adds = defaultdict(set)
                for v in views:
                    bucket = v["start"].replace(
                        second=0, microsecond=0, minute=v["start"].minute // 5 * 5
                    )
                    while bucket < end and (not v["end"] or bucket < v["end"]):
                        adds[bucket].add(v["user"])
                        bucket += timedelta(minutes=5)

                t = gds.replace(second=0, microsecond=0, minute=gds.minute // 5 * 5)
                while t <= gde:
                    ws.append(
                        [
                            t.astimezone(tz).date(),
                            t.astimezone(tz).time(),
                            len(adds[t]) if t in adds else "",
                        ]
                    )
                    t += timedelta(minutes=5)

                day += timedelta(days=1)

    wb.save(io)
    io.seek(0)

    sf = StoredFile.objects.create(
        world=world,
        date=now(),
        filename="report.xlsx",
        expires=now() + timedelta(hours=2),
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        public=True,
        user=None,
    )
    sf.file.save("report.xlsx", ContentFile(io.read()))
    return sf.file.url


@app.task(base=WorldTask)
def generate_session_views(world, input=None):
    wb = Workbook(write_only=True)
    io = BytesIO()
    tz = pytz.timezone(world.timezone)
    begin = dateutil.parser.parse(input.get("begin"))
    if is_naive(begin):
        make_aware(begin, tz)
    begin = begin.astimezone(tz)
    end = dateutil.parser.parse(input.get("end"))
    if is_naive(end):
        make_aware(end, tz)
    end = end.astimezone(tz)
    ws = wb.create_sheet("Sessions")
    ws.append(
        [
            "Title",
            "Room",
            "Start date",
            "Start time",
            "End date",
            "End time",
            "Viewership (approx. unique users)",
        ]
    )

    room_cache = {r.pretalx_id: r for r in world.rooms.filter(deleted=False)}
    schedule = get_schedule(world, fail_silently=False)
    for talk in schedule.get("talks", []):
        talk_start = dateutil.parser.parse(talk["start"])
        talk_end = dateutil.parser.parse(talk["end"])
        if talk_start > end or talk_end < begin:
            continue

        try:
            viewers = (
                RoomView.objects.filter(room__pretalx_id=talk["room"])
                .exclude(Q(end__lt=talk_start) | Q(start__gt=talk_end))
                .values("user")
                .distinct()
                .count()
            )

            ws.append(
                [
                    pretalx_uni18n(talk["title"]),
                    room_cache[talk["room"]].name
                    if talk["room"] in room_cache
                    else "?",
                    talk_start.astimezone(tz).date(),
                    talk_start.astimezone(tz).time(),
                    talk_end.astimezone(tz).date(),
                    talk_end.astimezone(tz).time(),
                    viewers,
                ]
            )
        except ValueError:  # e.g. pretalx_id not numeric
            pass

    wb.save(io)
    io.seek(0)

    sf = StoredFile.objects.create(
        world=world,
        date=now(),
        filename="report.xlsx",
        expires=now() + timedelta(hours=2),
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        public=True,
        user=None,
    )
    sf.file.save("report.xlsx", ContentFile(io.read()))
    return sf.file.url


@app.task(base=WorldTask)
def generate_views(world, input=None):
    wb = Workbook(write_only=True)
    io = BytesIO()
    tz = pytz.timezone(world.timezone)
    begin = dateutil.parser.parse(input.get("begin"))
    if is_naive(begin):
        make_aware(begin, tz)
    begin = begin.astimezone(tz)
    end = dateutil.parser.parse(input.get("end"))
    if is_naive(end):
        make_aware(end, tz)
    end = end.astimezone(tz)

    ws = wb.create_sheet("Room views")
    header = [
        "Room",
        "Start",
        "End",
        "User ID",
        "External ID",
        "User name",
    ]
    for n in world.config.get("profile_fields", []):
        header.append(n.get("label") or "")
    ws.append(header)
    rvq = (
        RoomView.objects.filter(
            Q(end__isnull=True) | Q(end__gte=begin),
            start__lte=end,
            room__world=world,
        )
        .select_related("room", "user")
        .order_by("start")
    )
    for v in rvq:
        u = v.user
        if u.profile.get("display_name"):
            ws.append(
                [
                    v.room.name,
                    v.start.astimezone(pytz.timezone(world.timezone)).strftime(
                        "%d.%m.%Y %H:%M:%S"
                    ),
                    (v.end or now())
                    .astimezone(pytz.timezone(world.timezone))
                    .strftime("%d.%m.%Y %H:%M:%S"),
                    str(u.pk),
                    u.token_id,
                    u.profile.get("display_name"),
                ]
                + [
                    (u.profile["fields"].get(n.get("id"), "") or "").strip()
                    for n in world.config.get("profile_fields", [])
                ]
            )

    ws = wb.create_sheet("World views")
    header = [
        "Start",
        "End",
        "User ID",
        "External ID",
        "User name",
    ]
    for n in world.config.get("profile_fields", []):
        header.append(n.get("label") or "")
    ws.append(header)
    rvq = (
        WorldView.objects.filter(
            Q(end__isnull=True) | Q(end__gte=begin),
            start__lte=end,
            world=world,
        )
        .select_related("user")
        .order_by("start")
    )
    for v in rvq:
        u = v.user
        if u.profile.get("display_name"):
            ws.append(
                [
                    v.start.astimezone(pytz.timezone(world.timezone)).strftime(
                        "%d.%m.%Y %H:%M:%S"
                    ),
                    (v.end or now())
                    .astimezone(pytz.timezone(world.timezone))
                    .strftime("%d.%m.%Y %H:%M:%S"),
                    str(u.pk),
                    u.token_id,
                    u.profile.get("display_name"),
                ]
                + [
                    (u.profile["fields"].get(n.get("id"), "") or "").strip()
                    for n in world.config.get("profile_fields", [])
                ]
            )

    ws = wb.create_sheet("Exhibitor views")
    header = [
        "Exhibition room",
        "Exhibitor",
        "Datetime",
        "User ID",
        "External ID",
        "User name",
    ]
    for n in world.config.get("profile_fields", []):
        header.append(n.get("label") or "")
    ws.append(header)
    rvq = (
        ExhibitorView.objects.filter(
            datetime__gte=begin,
            datetime__lte=end,
            exhibitor__world=world,
        )
        .select_related("exhibitor__room", "exhibitor", "user")
        .order_by("datetime")
    )
    for v in rvq:
        u = v.user
        if u.profile.get("display_name"):
            ws.append(
                [
                    v.exhibitor.room.name,
                    v.exhibitor.name,
                    v.datetime.astimezone(pytz.timezone(world.timezone)).strftime(
                        "%d.%m.%Y %H:%M:%S"
                    ),
                    str(u.pk),
                    u.token_id,
                    u.profile.get("display_name"),
                ]
                + [
                    (u.profile["fields"].get(n.get("id"), "") or "").strip()
                    for n in world.config.get("profile_fields", [])
                ]
            )

    wb.save(io)
    io.seek(0)

    sf = StoredFile.objects.create(
        world=world,
        date=now(),
        filename="report.xlsx",
        expires=now() + timedelta(hours=2),
        type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        public=True,
        user=None,
    )
    sf.file.save("report.xlsx", ContentFile(io.read()))
    return sf.file.url
